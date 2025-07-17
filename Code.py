import openpyxl
import os
import re

# Templates for C code generation
# Template for the main function
main_template = """#define ON 1
#define OFF 0
#define unsigned short int WORD
{variable_declarations}
{init_function}
{interface_function}
int main() {{
    WORD dopNos[10];
    switch(event) {{
{cases}
        default:
            break;
    }}
}}
"""

# Template for the init function
init_function_template = """void init() {{
{init_variables}}}
"""

# Template for initialization code inside the init function
init_variable_template = "    {name} = {initial_value};\n"

# Template for each function call
function_calling_template = """
                {dop_initialization}
                func({BusNo}, {SubSystem}, dopNos, {count}, {Status});
"""

# Template for each case in the switch statement
case_template = """
        case {event}:
            {variables}
            if {cond} {{
                {function_calling}
                {t_variable} = time;
                event = {next_event};
            }}
            break;
"""

# Template for variable declarations inside the main function
variable_declarations_template = """
int event;
float time;
{variables}"""

# Template for array initializations inside the if condition
array_initialization_template = "dopNos[{i}] = \"{value}\";\n\t\t\t\t"

# Template for variable declarations
variable_template = "{type} {name};\n"

# Template for the interface function
interface_function_template = """void interface() {{
{interface_assignments}}}
"""


def map_headers_to_columns(sheet):
    headers = [sheet.cell(row=1, column=col).value.lower() for col in range(1, sheet.max_column + 1)]
    header_map = {}

    for header_index, header_name in enumerate(headers, start=1):
        if header_name in ['subsystem', 'subsys']:
            header_map['subsys'] = header_index
        elif header_name in ['condition', 'cond']:
            header_map['cond'] = header_index
        elif header_name in ['operator', 'operation', 'oprtn']:
            header_map['oprtn'] = header_index
        elif header_name == 'busno':
            header_map['busno'] = header_index
        elif header_name == 'dopdata':
            header_map['dopdata'] = header_index
        elif header_name == 'status':
            header_map['status'] = header_index
        elif header_name == 'variable':
            header_map['variable'] = header_index
        elif header_name in ['func_to_call']:
            header_map['func_to_call'] = header_index
        elif header_name == 'event':
            header_map['event'] = header_index

    return header_map


def check_column_order(sheet):
    header_map = map_headers_to_columns(sheet)
    required_order = ['event', 'busno', 'subsys', 'dopdata', 'status', 'variable', 'func_to_call', 'cond', 'oprtn']

    for i, header in enumerate(required_order):
        if header.lower() not in header_map or header_map[header.lower()] != i + 1:
            raise ValueError(f"Columns are not in the correct order. Expected order: {', '.join(required_order)}")


def check_event_column(sheet):
    events = []
    header_map = map_headers_to_columns(sheet)
    required_columns = ['event', 'busno', 'dopdata', 'subsys', 'status']

    for row in range(2, sheet.max_row + 1):
        for column in required_columns:
            value = sheet.cell(row=row, column=header_map[column]).value
            if value is None:
                raise ValueError(f"Missing {column} value in row {row}.")

        event = sheet.cell(row=row, column=header_map['event']).value
        if event in events:
            raise ValueError(f"Duplicate event value '{event}' found in row {row}.")
        events.append(event)

    if events != list(range(1, len(events) + 1)):
        raise ValueError("Event values are not in proper order or there are missing values.")


def process_conditions(full_condition):
    conditions = []
    operators = []

    for i in range(0, len(full_condition), 2):
        if i < len(full_condition):
            conditions.append(f'({full_condition[i]})')
    if len(full_condition) > 1:
        operators = [full_condition[i] for i in range(1, len(full_condition), 2) if i < len(full_condition)]

    if not conditions:
        return "", [], []  # Return empty strings and lists if no conditions found

    result = conditions[0]
    for condition, operator in zip(conditions[1:], operators):
        result += f" {operator} {condition}"
    # Replace logical operators with C-style logical operators
    result = result.replace("and", "&&").replace("or", "||")
    return result, conditions, operators

def generate_interface_assignments(xyz):
    assignments = ""
    for i, item in enumerate(xyz):
        assignments += f"    {item} = ;\n"  # Adjust as needed for initialization values
    return assignments

def strip_conditions_up_to_operator(conditions):
    # Define a list of complex operators to check for
    complex_operators = ["+=", "-=", "*=", "/=", "%=", "<=", ">=", "==", "!="]

    # List of logical operators to exclude
    logical_operators = ["or", "and"]

    # Define regex pattern to match valid identifiers (alphabetic characters or digits with special characters like '_')
    pattern = r'\b\w+\b'

    # List to store extracted elements
    xyz = []

    for condition in conditions:
        # Initialize the split index to None
        split_index = None
        operator_found = False

        # Check for complex operators first
        for op in complex_operators:
            if op in condition:
                split_index = condition.index(op)
                operator_found = True
                break

        # If no complex operator found, check for single character operators
        if not operator_found:
            for op in ["-", "%", "=", "<", ">", "+", "*", "/"]:
                if op in condition:
                    split_index = condition.index(op)
                    break

        if split_index is not None:
            # Extract the part before and after the first operator
            part_before_operator = condition[:split_index].strip()
            part_after_operator = condition[split_index + len(condition[split_index:].split(' ')[0]):].strip()

            # Use regex to extract individual variables or terms
            terms_before = re.findall(pattern, part_before_operator)
            terms_after = re.findall(pattern, part_after_operator)

            # Add terms before operator to the list if they are not numeric
            for term in terms_before:
                if term and term not in xyz and term not in logical_operators and not term.isdigit():
                    xyz.append(term)

            # Add terms after operator to the list if they are not numeric
            for term in terms_after:
                if term and term not in xyz and term not in logical_operators and not term.isdigit():
                    xyz.append(term)
        else:
            # If no operator is found, include the whole condition if it's not empty
            terms = re.findall(pattern, condition.strip())
            for term in terms:
                if term and term not in logical_operators and not term.isdigit():
                    xyz.append(term)

    # Remove duplicates
    xyz = list(set(xyz))

    return xyz

def generate_cases(sheet):
    global min_index
    cases = ""
    xyz = []
    seen_items = set()
    header_map = map_headers_to_columns(sheet)
    max_row = sheet.max_row
    max_star_counts = []  # Initialize list to store max '*' counts for each row

    for row in range(2, max_row + 1):  # Assuming the first row is headers
        event = sheet.cell(row=row, column=header_map['event']).value
        condition = sheet.cell(row=row, column=header_map.get('cond', header_map.get('condition'))).value
        operation = sheet.cell(row=row, column=header_map.get('oprtn', header_map.get('operator', header_map.get('operation')))).value
        BusNo = sheet.cell(row=row, column=header_map['busno']).value
        SubSystem = sheet.cell(row=row, column=header_map.get('subsys', header_map.get('subsystem'))).value
        Status = sheet.cell(row=row, column=header_map['status']).value
        dopData = sheet.cell(row=row, column=header_map['dopdata']).value.split(',') if sheet.cell(row=row, column=header_map['dopdata']).value else []
        variables = sheet.cell(row=row, column=header_map['variable']).value.strip() if sheet.cell(row=row, column=header_map['variable']).value else ""

        dop_count = len(dopData)

        # Collect all values from cells after cond and oprtn columns
        additional_conditions = []
        column_start_index = header_map['oprtn'] + 1 if 'oprtn' in header_map else header_map['operation'] + 1 if 'operation' in header_map else None

        if column_start_index:
            while column_start_index <= sheet.max_column:
                cell_value = sheet.cell(row=row, column=column_start_index).value
                if cell_value is None or cell_value.strip() == "":
                    break
                additional_conditions.append(cell_value.strip())
                column_start_index += 1

        # Concatenate cond, oprtn, and additional conditions into full_condition
        full_condition = []
        if condition and condition.strip() != "":
            full_condition.append(condition.strip())
        elif condition is None:  # Handle None type explicitly
            full_condition.append('1')  # Default to '1' if cond column is empty

        if operation and operation.strip() != "":
            full_condition.append(operation.strip())
        full_condition.extend(additional_conditions)

        # Process conditions using process_conditions function
        processed_condition, conditions, operators = process_conditions(full_condition)

        # Clear the full_condition list before populating it with interleaved conditions and operators
        full_condition = []

        # Interleave conditions and operators
        max_len = max(len(conditions), len(operators))
        for i in range(max_len):
            if i < len(conditions):
                full_condition.append(conditions[i])
            if i < len(operators):
                full_condition.append(operators[i])

        for item in strip_conditions_up_to_operator(full_condition):
            if item and item not in seen_items:
                xyz.append(item)
                seen_items.add(item)

        # Handle '*' priority operators if applicable
        if len(operators) > 1:
            # Extract operators
            operators = [op.strip() for op in full_condition[1::2]]
            # Check if priorities are correctly specified
            if any(not op.startswith('*') for op in operators):
                raise ValueError("Make sure the priorities are specified correctly.")

            # Logic to handle '*' priority operators
            star_counts = [op.count('*') for op in operators]
            max_star_count = max(star_counts) if star_counts else 0
            max_star_counts.append(max_star_count)

            # Ensure no adjacent operators with the same star count
            for i in range(1, len(star_counts)):
                if star_counts[i] == star_counts[i - 1]:
                    raise ValueError(f"Adjacent operators with the same number of '*'s found at row {row}")

            # Add parentheses based on star count
            for i in range(1, max_star_count + 1):
                while True:
                    min_star_count = float('inf')
                    min_operator = None
                    min_index = None  # Initialize min_index
                    found_operator = False

                    for op in operators:
                        star_count = len(op) - len(op.lstrip('*'))
                        if star_count == i and star_count < min_star_count:
                            min_star_count = star_count
                            min_operator = op
                            min_index = full_condition.index(min_operator)  # Set min_index
                            found_operator = True

                    if not found_operator:
                        break

                    # Remove leading '*'
                    min_operator = min_operator.lstrip('*')
                    full_condition[min_index] = min_operator

                    # Insert parentheses
                    if min_index == 0 or min_index == 1:
                        full_condition.insert(0, "(")
                        full_condition.insert(min_index + 3, ")")
                    elif min_index > 1:
                        full_condition.insert(min_index - 1, "(")
                        full_condition.insert(min_index + 3, ")")
                    else:
                        raise ValueError("Invalid Index!!")

                    # Merge the condition within the parentheses
                    start_index = full_condition.index('(')
                    end_index = full_condition.index(')', start_index)
                    merged_element = ' '.join(full_condition[start_index:end_index + 1])
                    full_condition = full_condition[:start_index] + [merged_element] + full_condition[end_index + 1:]

                    # Update the operators list
                    operators = [op.strip() for op in full_condition[1::2]]

        # Generate initialization and function calling code
        dop_initialization = "".join([array_initialization_template.format(i=i, value=dop.strip()) for i, dop in enumerate(dopData)])
        function_calling = function_calling_template.format(
            dop_initialization=dop_initialization,
            BusNo=BusNo,
            count=dop_count,
            SubSystem=SubSystem,
            Status=Status
        )

        t_variable = f"timeE{event}"
        next_event = event + 1

        # Prepare variables code
        if variables and variables.lower() not in ['none', 'null']:
            variable_initializations = ""
            for var in variables.split(';'):
                if var.strip() and '=' in var:
                    variable_initializations += f"    {var.strip()};\n"
            variables_code = variable_initializations
        else:
            variables_code = ""

        # Format the final condition and add to cases
        updated_full_condition = ' '.join(full_condition).replace("and", "&&").replace("or", "||")
        cases += case_template.format(event=event, variables=variables_code, cond=updated_full_condition,
                                      function_calling=function_calling,
                                      t_variable=t_variable, next_event=next_event)
    return cases, xyz
def generate_variables_and_inits(sheet, xyz):
    variables = ""
    inits = "    event = 1;\n    time = 0.0;\n"
    header_map = map_headers_to_columns(sheet)
    max_row = sheet.max_row
    declared_vars = set()
    initialized_vars = set()

    for row in range(2, max_row + 1):
        event = sheet.cell(row=row, column=header_map['event']).value
        variable_name = f"timeE{event}"
        variables += variable_template.format(type="float", name=variable_name)
        declared_vars.add(variable_name)  # Track declared variables
        initialized_vars.add(variable_name)  # Track initialized variables
        inits += init_variable_template.format(name=variable_name, initial_value="0.0")

    # Add a comment and xyz variables to variables string
    variables += "\n// other variables\n"
    new_vars = ""
    new_inits = ""
    for item in xyz:
        if item not in declared_vars:  # Add only if not already declared
            new_vars += variable_template.format(type="float", name=item)
            declared_vars.add(item)  # Track new declarations
        if item not in initialized_vars:  # Add only if not already initialized
            new_inits += init_variable_template.format(name=item, initial_value="0.0")
            initialized_vars.add(item)  # Track new initializations

    # Add the xyz variables declarations and initialization to variables string with a comment
    if new_vars:
        variables += new_vars
    if new_inits:
        inits += "\n    // other variables initialization\n" + new_inits

    return variables, inits
def generate_c_code(sheet):
    check_column_order(sheet)
    check_event_column(sheet)
    cases, xyz = generate_cases(sheet)  # Capture xyz from generate_cases
    variables, init_variables = generate_variables_and_inits(sheet, xyz)  # Pass xyz to generate_variables_and_inits
    init_function = init_function_template.format(init_variables=init_variables)
    interface_assignments = generate_interface_assignments(xyz)
    interface_function = interface_function_template.format(interface_assignments=interface_assignments)
    variable_declarations = variable_declarations_template.format(variables=variables)
    return main_template.format(init_function=init_function, variable_declarations=variable_declarations,
                                interface_function=interface_function, cases=cases)

def save_c_code(filename, c_code):
    with open(filename, 'w') as file:
        file.write(c_code)


FileName = input("Enter the name of the file: ")
if FileName.endswith(".xlsx"):
    FileName = FileName
else:
    FileName = FileName + ".xlsx"

if os.path.isfile(FileName):
    def get_sheet(WorkBook, SheetName=None):
        sheets = WorkBook.sheetnames
        if len(sheets) == 1:
            return WorkBook[sheets[0]]
        else:
            if SheetName is None:
                print("Available sheets:", sheets)
                while True:
                    SheetName = input("Please enter the name of the sheet to select: ")
                    if SheetName in sheets:
                        break
                    else:
                        print(f"'{SheetName}' is not a valid sheet name. Please try again.")
            return WorkBook[SheetName]


    try:
        WorkBook = openpyxl.load_workbook(FileName)
        SelectedSheet = get_sheet(WorkBook)
        header_map = map_headers_to_columns(SelectedSheet)

        try:
            c_code = generate_c_code(SelectedSheet)
            output_filename = "output.c"
            save_c_code(output_filename, c_code)
            print(f"C code has been generated and saved to {output_filename}")
        except ValueError as e:
            print(e)
    except PermissionError:
        print(
            f"Permission denied: '{FileName}'. Please check if the file is open in another application or if you have the necessary permissions.")
else:
    print("File not found.")
